"""
توليد ملف Excel لكشف الرواتب الشهري
=====================================
يُنشئ ملفاً واحداً يحتوي على رواتب جميع الموظفين للفترة المحددة.
يستخدم openpyxl.
"""
import io
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

_MONTHS_AR = {
    1:'يناير', 2:'فبراير', 3:'مارس', 4:'أبريل',
    5:'مايو',  6:'يونيو',  7:'يوليو', 8:'أغسطس',
    9:'سبتمبر', 10:'أكتوبر', 11:'نوفمبر', 12:'ديسمبر',
}

STATUS_MAP = {
    'draft':    'مسودة',
    'pending':  'بانتظار الاعتماد',
    'approved': 'معتمد',
    'paid':     'مصروف',
    'cancelled':'ملغي',
}

# ── ألوان ───────────────────────────────────────
CLR_HEADER   = '1a1d2e'   # رأس الجدول - داكن
CLR_TOTAL    = '28a745'   # صف المجاميع - أخضر
CLR_ALT      = 'f0f4ff'   # صفوف متناوبة
CLR_TITLE    = '1a1d2e'
CLR_WHITE    = 'FFFFFF'

NUM_FMT = '#,##0.00'

thin = Side(style='thin', color='DEE2E6')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, color='000000', size=10):
    return Font(bold=bold, color=color, size=size, name='Arial')


def _align(horizontal='center', vertical='center', wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical,
                     wrap_text=wrap, readingOrder=2)


def export_payroll_excel(payrolls, period):
    """
    إنشاء HttpResponse يحتوي على Excel كشف الرواتب.

    payrolls : QuerySet أو list من Payroll
    period   : PayrollPeriod أو أي كائن له month و year
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'كشف الرواتب'
    ws.sheet_view.rightToLeft = True   # RTL

    month_label = _MONTHS_AR.get(period.month, str(period.month))
    period_str  = f"{month_label} {period.year}"

    # ══════════════════════════════════════════════
    # صف العنوان الرئيسي  (Row 1)
    # ══════════════════════════════════════════════
    COLS = 15
    ws.merge_cells(f'A1:{get_column_letter(COLS)}1')
    title_cell = ws['A1']
    title_cell.value = f'كشف الرواتب الشهري  —  {period_str}'
    title_cell.font      = _font(bold=True, color=CLR_WHITE, size=14)
    title_cell.fill      = _fill(CLR_HEADER)
    title_cell.alignment = _align()

    # صف الفرعي (Row 2)
    ws.merge_cells(f'A2:{get_column_letter(COLS)}2')
    sub_cell = ws['A2']
    sub_cell.value = (
        f'تاريخ الطباعة: {datetime.now().strftime("%Y-%m-%d %H:%M")}    |'
        f'    عدد الموظفين: {len(payrolls)}'
    )
    sub_cell.font      = _font(color='6c757d', size=10)
    sub_cell.fill      = _fill('f8f9fa')
    sub_cell.alignment = _align()

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # ══════════════════════════════════════════════
    # رأس الجدول (Row 3)
    # ══════════════════════════════════════════════
    headers = [
        '#', 'اسم الموظف', 'رقم الموظف', 'الفرع',
        'الأساسي', 'سكن', 'مواصلات', 'بدلات أخرى',
        'إضافي', 'تأمين', 'غياب', 'سلفة', 'خصومات', 'الصافي', 'الحالة',
    ]
    col_widths = [5, 26, 14, 16, 14, 12, 12, 12, 12, 12, 12, 12, 12, 14, 14]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = _font(bold=True, color=CLR_WHITE, size=10)
        cell.fill      = _fill(CLR_HEADER)
        cell.alignment = _align()
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 22

    # ══════════════════════════════════════════════
    # صفوف البيانات
    # ══════════════════════════════════════════════
    totals = {k: Decimal('0') for k in [
        'basic','housing','transport','other',
        'overtime','ins','absence','advance','deductions','net',
    ]}

    NUMBER_COLS = list(range(5, 14 + 1))   # أعمدة 5-14 هي أرقام

    for i, p in enumerate(payrolls, 1):
        row = 3 + i
        row_fill = _fill(CLR_ALT) if i % 2 == 0 else _fill(CLR_WHITE)

        values = [
            i,
            str(p.employee.employee_name if p.employee else ''),
            str(p.employee.employee_number if p.employee else ''),
            str(p.employee.branch or ''),
            p.basic_salary         or Decimal('0'),
            p.housing_allowance    or Decimal('0'),
            p.transport_allowance  or Decimal('0'),
            p.other_allowances     or Decimal('0'),
            p.overtime_amount      or Decimal('0'),
            p.social_insurance     or Decimal('0'),
            p.absence_deduction    or Decimal('0'),
            p.advance_deduction    or Decimal('0'),
            p.deductions           or Decimal('0'),
            p.net_salary           or Decimal('0'),
            STATUS_MAP.get(p.status, p.status),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=float(val) if isinstance(val, Decimal) else val)
            cell.fill      = row_fill
            cell.border    = BORDER
            cell.alignment = _align(horizontal='center')
            cell.font      = _font(size=10)
            if col_idx in NUMBER_COLS:
                cell.number_format = NUM_FMT

        # تراكم المجاميع
        totals['basic']      += p.basic_salary         or Decimal('0')
        totals['housing']    += p.housing_allowance    or Decimal('0')
        totals['transport']  += p.transport_allowance  or Decimal('0')
        totals['other']      += p.other_allowances     or Decimal('0')
        totals['overtime']   += p.overtime_amount      or Decimal('0')
        totals['ins']        += p.social_insurance     or Decimal('0')
        totals['absence']    += p.absence_deduction    or Decimal('0')
        totals['advance']    += p.advance_deduction    or Decimal('0')
        totals['deductions'] += p.deductions           or Decimal('0')
        totals['net']        += p.net_salary           or Decimal('0')

        ws.row_dimensions[row].height = 18

    # ══════════════════════════════════════════════
    # صف المجاميع
    # ══════════════════════════════════════════════
    total_row = 3 + len(payrolls) + 1
    total_values = [
        '', 'الإجمالي', '', '',
        float(totals['basic']),
        float(totals['housing']),
        float(totals['transport']),
        float(totals['other']),
        float(totals['overtime']),
        float(totals['ins']),
        float(totals['absence']),
        float(totals['advance']),
        float(totals['deductions']),
        float(totals['net']),
        '',
    ]
    for col_idx, val in enumerate(total_values, 1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font      = _font(bold=True, color=CLR_WHITE, size=11)
        cell.fill      = _fill(CLR_TOTAL)
        cell.alignment = _align()
        cell.border    = BORDER
        if col_idx in NUMBER_COLS and isinstance(val, float):
            cell.number_format = NUM_FMT

    ws.row_dimensions[total_row].height = 22

    # ══════════════════════════════════════════════
    # تجميد الرأس
    # ══════════════════════════════════════════════
    ws.freeze_panes = 'A4'

    # ══════════════════════════════════════════════
    # إعداد الطباعة
    # ══════════════════════════════════════════════
    ws.page_setup.orientation     = 'landscape'
    ws.page_setup.paperSize       = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage       = True
    ws.page_setup.fitToWidth      = 1
    ws.page_setup.fitToHeight     = 0
    ws.print_title_rows           = '1:3'

    # ══════════════════════════════════════════════
    # إرجاع الاستجابة
    # ══════════════════════════════════════════════
    buffer = io.BytesIO()
    wb.save(buffer)

    filename = f"payroll_{period.year}_{period.month:02d}.xlsx"
    
    from django.core.files.base import ContentFile
    if hasattr(period, 'excel_file'):
        period.excel_file.save(filename, ContentFile(buffer.getvalue()), save=True)

    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
